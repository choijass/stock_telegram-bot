# Converted from Google Colab notebook
# Source notebook: 6.etf_20일선_ 종합버젼

# %% [code] cell 1
# =========================================================
# [최종 통합본 - 시장요약 강화 + 구글시트 셀 한도 오류 수정]
# RS + MA + 거래대금 증가율 + 변동성조정모멘텀 + 시장필터 + 신규진입
# =========================================================


import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime
from zoneinfo import ZoneInfo

# ---------------------------------------------------------
# 0) 기본 설정
# ---------------------------------------------------------
KST = ZoneInfo("Asia/Seoul")
NOW = datetime.now(KST)
TODAY = NOW.strftime("%Y-%m-%d")
CURRENT_YEAR = NOW.year

YEARS_BACK = 10
MA_LIST = [5, 10, 20, 200]
FORWARD_DAYS = [5, 10, 20, 60]

SIGNAL_COOLDOWN = 20
MIN_TURNOVER = 5_000_000_000   # 20일 평균 거래대금 50억 이상

OUT_XLSX = "/content/korea_universe_with_ra_momentum_final_fixed_v2.xlsx"
GSHEET_NAME = "Korea_Universe_RA_Momentum_Final_Fixed_V2"

# RS 가중치
RS_WEIGHT_20 = 0.35
RS_WEIGHT_60 = 0.35
RS_WEIGHT_120 = 0.20
RS_WEIGHT_240 = 0.10

# 최종 점수 가중치
WEIGHT_RS = 0.35
WEIGHT_MA = 0.25
WEIGHT_TURNOVER = 0.15
WEIGHT_RA = 0.25

# 시장 필터 점수
MARKET_BONUS_STRONG = 10
MARKET_BONUS_NEUTRAL = 3
MARKET_PENALTY_WEAK = -12
MARKET_PENALTY_VERY_WEAK = -20

STRICT_NEW_ENTRY = False

# 변동성 조정 모멘텀 설정
RA_1M = 21
RA_3M = 63
RA_6M = 126
RA_LAST_N_DAYS = 10

# ---------------------------------------------------------
# 1) 분석 대상
# ---------------------------------------------------------
INDEX_TICKERS = {
    "KOSPI": "^KS11",
    "KOSDAQ": "^KQ11",
}

UNIVERSE_TICKERS = {
    "091170.KS": "KODEX 은행",
    "140700.KS": "KODEX 보험",
    "102970.KS": "KODEX 증권",
    "117700.KS": "KODEX 건설",
    "300950.KS": "KODEX 게임산업",
    "395160.KS": "KODEX 시스템반도체",
    "445290.KS": "KODEX K-로봇 액티브",
    "117460.KS": "KODEX 에너지화학",
    "091160.KS": "KODEX 반도체",
    "000660.KS": "SK하이닉스",
    "196170.KQ": "알테오젠",
    "244580.KS": "KODEX 바이오",
    "228800.KS": "TIGER 여행레저",
    "364970.KS": "TIGER 바이오 TOP 10",
    "091180.KS": "KODEX 자동차",
    "305540.KS": "TIGER 2차전지테마",
    "462010.KS": "TIGER 2차전지소재FN",
    "266360.KS": "KODEX 미디어&엔터테인먼트",
    "395150.KS": "KODEX 웹툰&드라마",
    "367760.KS": "RISE 5G테크",
    "228790.KS": "TIGER 화장품",
    "463250.KS": "TIGER 우주방산",
    "157490.KS": "TIGER 소프트웨어",
    "449450.KS": "PLUS K 방산",
    "139230.KS": "TIGER 200 중공업",
    "150460.KS": "TIGER 중국소비테마",
    "139280.KS": "TIGER 경기방어",
    "438900.KS": "HANARO FN K-푸드",
    "381570.KS": "HANARO FN친환경에너지",
    "210780.KS": "KODEX 코스피고배당",
    "466920.KS": "SOL 조선TOP3플러스",
    "475300.KS": "SOL 반도체전공정",
    "475310.KS": "SOL 반도체후공정",
    "307510.KS": "TIGER 의료기기",
    "433500.KS": "ACE 원자력테마딥서치",
    "483020.KS": "KIWOOM 의료AI",
    "0000J0.KS": "PLUS 한화그룹주",
    "0008T0.KS": "SOL 화장품TOP3플러스",
    "495040.KS": "PLUS 코리아밸류업",
    "496770.KS": "PLUS 글로벌방산",
    "395290.KS": "HANARO Fn K-POP&미디어",
    "102960.KS": "KODEX 기계장비",
    "307520.KS": "TIGER 지주회사",
    "365000.KS": "TIGER 인터넷TOP10",
    "463050.KS": "TIMEFOLIO K바이오액티브",
    "421320.KS": "PLUS 우주항공&UAM",
    "466930.KS": "SOL 자동차TOP3플러스",
    "457990.KS": "PLUS 태양광&ESS",
}

TARGETS = []
for k, v in INDEX_TICKERS.items():
    TARGETS.append({"name": k, "ticker": v, "asset_type": "지수"})
for ticker, name in UNIVERSE_TICKERS.items():
    TARGETS.append({"name": name, "ticker": ticker, "asset_type": "유니버스"})

# ---------------------------------------------------------
# 2) 유틸
# ---------------------------------------------------------
def flatten_columns(df):
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
    return df

def safe_round(v, digits=2):
    if pd.isna(v):
        return np.nan
    return round(float(v), digits)

def safe_mean(series):
    s = pd.Series(series).dropna()
    return np.nan if len(s) == 0 else s.mean()

def safe_median(series):
    s = pd.Series(series).dropna()
    return np.nan if len(s) == 0 else s.median()

def safe_winrate(series):
    s = pd.Series(series).dropna()
    return np.nan if len(s) == 0 else (s > 0).mean()

def percentile_score(series):
    s = pd.Series(series)
    return s.rank(method="average", pct=True) * 100

def clip_sheet_title(text, max_len=31):
    text = str(text).replace("/", "_").replace("\\", "_").replace("?", "_").replace("*", "_").replace("[", "").replace("]", "")
    return text[:max_len]

def classify_signal(diff20):
    if pd.isna(diff20):
        return "판단불가"
    if diff20 >= 2.0:
        return "상향돌파 매우 우세"
    elif diff20 >= 0.5:
        return "상향돌파 우세"
    elif diff20 <= -2.0:
        return "하향이탈 후 반등 매우 우세"
    elif diff20 <= -0.5:
        return "하향이탈 후 반등 우세"
    return "차이 작음"

def get_above_below_text(close_price, ma_value):
    if pd.isna(close_price) or pd.isna(ma_value):
        return "N/A"
    return "위" if close_price > ma_value else "아래"

def get_stack_status_from_values(ma5, ma10, ma20, ma200):
    if pd.isna(ma5) or pd.isna(ma10) or pd.isna(ma20) or pd.isna(ma200):
        return "판단불가"
    if ma5 > ma10 > ma20 > ma200:
        return "정배열"
    elif ma5 < ma10 < ma20 < ma200:
        return "역배열"
    return "혼조"

# ---------------------------------------------------------
# 3) 데이터 다운로드
# ---------------------------------------------------------
def download_price(ticker, start_date, end_date):
    try:
        df = yf.download(
            ticker,
            start=start_date,
            end=end_date,
            auto_adjust=True,
            progress=False
        )
        if df is None or df.empty:
            return pd.DataFrame()

        df = flatten_columns(df)
        use_cols = [c for c in ["Open", "High", "Low", "Close", "Volume"] if c in df.columns]
        df = df[use_cols].copy()

        df.index = pd.to_datetime(df.index)
        try:
            df.index = df.index.tz_localize(None)
        except Exception:
            pass

        df = df[~df.index.duplicated(keep="first")].sort_index()

        if "Volume" not in df.columns:
            df["Volume"] = np.nan

        return df
    except Exception as e:
        print(f"[다운로드 실패] {ticker}: {e}")
        return pd.DataFrame()

# ---------------------------------------------------------
# 4) 거래대금
# ---------------------------------------------------------
def add_turnover(df):
    x = df.copy()
    x["Turnover"] = x["Close"] * x["Volume"] if "Volume" in x.columns else np.nan
    x["Turnover5"] = x["Turnover"].rolling(5).mean()
    x["Turnover20"] = x["Turnover"].rolling(20).mean()
    x["Turnover60"] = x["Turnover"].rolling(60).mean()
    x["Turnover_Ratio_5_20"] = x["Turnover5"] / x["Turnover20"]
    x["Turnover_Ratio_1_20"] = x["Turnover"] / x["Turnover20"]
    return x

# ---------------------------------------------------------
# 5) 중복신호 제거
# ---------------------------------------------------------
def apply_cooldown(signal_series, cooldown=20):
    signal_series = signal_series.fillna(False).astype(bool)
    result = pd.Series(False, index=signal_series.index)
    last_idx = -10**9
    for i, val in enumerate(signal_series.values):
        if val and (i - last_idx > cooldown):
            result.iloc[i] = True
            last_idx = i
    return result

# ---------------------------------------------------------
# 6) 이벤트 데이터프레임
# ---------------------------------------------------------
def build_event_dataframe(df, ma_list=None, forward_days=None, cooldown=20, min_turnover=None):
    if ma_list is None:
        ma_list = MA_LIST
    if forward_days is None:
        forward_days = FORWARD_DAYS

    x = add_turnover(df.copy())
    x["Year"] = x.index.year
    x["Month"] = x.index.month

    for n in forward_days:
        x[f"FWD_{n}D"] = x["Close"].shift(-n) / x["Close"] - 1.0

    x["RET_20"] = x["Close"] / x["Close"].shift(20) - 1.0
    x["RET_60"] = x["Close"] / x["Close"].shift(60) - 1.0
    x["RET_120"] = x["Close"] / x["Close"].shift(120) - 1.0
    x["RET_240"] = x["Close"] / x["Close"].shift(240) - 1.0

    turnover_ok = pd.Series(True, index=x.index) if min_turnover is None else (x["Turnover20"] >= min_turnover)

    for ma in ma_list:
        x[f"MA{ma}"] = x["Close"].rolling(ma).mean()
        x[f"Prev_Close_{ma}"] = x["Close"].shift(1)
        x[f"Prev_MA{ma}"] = x[f"MA{ma}"].shift(1)

        raw_up = (
            (x[f"Prev_Close_{ma}"] <= x[f"Prev_MA{ma}"]) &
            (x["Close"] > x[f"MA{ma}"]) &
            turnover_ok
        )
        raw_down = (
            (x[f"Prev_Close_{ma}"] >= x[f"Prev_MA{ma}"]) &
            (x["Close"] < x[f"MA{ma}"]) &
            turnover_ok
        )

        x[f"UP_RAW_{ma}"] = raw_up
        x[f"DOWN_RAW_{ma}"] = raw_down
        x[f"UP_{ma}"] = apply_cooldown(raw_up, cooldown)
        x[f"DOWN_{ma}"] = apply_cooldown(raw_down, cooldown)
        x[f"DIST_MA{ma}"] = x["Close"] / x[f"MA{ma}"] - 1.0

    return x

# ---------------------------------------------------------
# 7) 백테스트 요약
# ---------------------------------------------------------
def summarize_total_by_ma(x, ma_list=None, forward_days=None):
    if ma_list is None:
        ma_list = MA_LIST
    if forward_days is None:
        forward_days = FORWARD_DAYS

    rows = []
    for ma in ma_list:
        row = {"이평선": f"MA{ma}", "MA": ma}
        up = x[x[f"UP_{ma}"]].copy()
        down = x[x[f"DOWN_{ma}"]].copy()

        row["상향돌파_표본수"] = len(up)
        row["하향이탈_표본수"] = len(down)
        row["상향돌파_RAW표본수"] = int(x[f"UP_RAW_{ma}"].sum())
        row["하향이탈_RAW표본수"] = int(x[f"DOWN_RAW_{ma}"].sum())

        for n in forward_days:
            up_vals = up[f"FWD_{n}D"].dropna()
            down_vals = down[f"FWD_{n}D"].dropna()

            up_mean = safe_mean(up_vals)
            down_mean = safe_mean(down_vals)

            row[f"상향돌파_{n}일평균수익률(%)"] = safe_round(up_mean * 100) if pd.notna(up_mean) else np.nan
            row[f"하향이탈_{n}일평균수익률(%)"] = safe_round(down_mean * 100) if pd.notna(down_mean) else np.nan
            row[f"상향돌파_{n}일중앙값수익률(%)"] = safe_round(safe_median(up_vals) * 100) if len(up_vals) else np.nan
            row[f"하향이탈_{n}일중앙값수익률(%)"] = safe_round(safe_median(down_vals) * 100) if len(down_vals) else np.nan
            row[f"상향돌파_{n}일승률(%)"] = safe_round(safe_winrate(up_vals) * 100) if len(up_vals) else np.nan
            row[f"하향이탈_{n}일승률(%)"] = safe_round(safe_winrate(down_vals) * 100) if len(down_vals) else np.nan
            row[f"상향-하향_{n}일평균차이(%p)"] = safe_round((up_mean - down_mean) * 100) if pd.notna(up_mean) and pd.notna(down_mean) else np.nan

        row["종합판단"] = classify_signal(row.get("상향-하향_20일평균차이(%p)", np.nan))
        rows.append(row)

    out = pd.DataFrame(rows)
    if not out.empty and "상향-하향_20일평균차이(%p)" in out.columns:
        out = out.sort_values("상향-하향_20일평균차이(%p)", ascending=False).reset_index(drop=True)
    return out

def summarize_yearly_by_ma(x, ma_list=None, forward_days=None):
    if ma_list is None:
        ma_list = MA_LIST
    if forward_days is None:
        forward_days = FORWARD_DAYS

    rows = []
    years = sorted([y for y in x["Year"].dropna().unique()])
    for ma in ma_list:
        for year in years:
            yy = x[x["Year"] == year].copy()
            up = yy[yy[f"UP_{ma}"]].copy()
            down = yy[yy[f"DOWN_{ma}"]].copy()

            row = {"Year": int(year), "이평선": f"MA{ma}", "MA": ma, "상향돌파_표본수": len(up), "하향이탈_표본수": len(down)}
            for n in forward_days:
                up_vals = up[f"FWD_{n}D"].dropna()
                down_vals = down[f"FWD_{n}D"].dropna()
                up_mean = safe_mean(up_vals)
                down_mean = safe_mean(down_vals)
                row[f"상향돌파_{n}일평균수익률(%)"] = safe_round(up_mean * 100) if pd.notna(up_mean) else np.nan
                row[f"하향이탈_{n}일평균수익률(%)"] = safe_round(down_mean * 100) if pd.notna(down_mean) else np.nan
                row[f"상향-하향_{n}일평균차이(%p)"] = safe_round((up_mean - down_mean) * 100) if pd.notna(up_mean) and pd.notna(down_mean) else np.nan
            rows.append(row)

    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["Year", "MA"], ascending=[True, True]).reset_index(drop=True)
    return out

def build_event_log(x, ma_list=None, forward_days=None):
    if ma_list is None:
        ma_list = MA_LIST
    if forward_days is None:
        forward_days = FORWARD_DAYS

    logs = []
    for ma in ma_list:
        cols = ["Close", f"MA{ma}", "Volume", "Turnover", "Turnover20", "Turnover_Ratio_5_20", "Year", "Month"] + [f"FWD_{n}D" for n in forward_days]

        up_log = x[x[f"UP_{ma}"]].copy()
        if not up_log.empty:
            temp = up_log[cols].copy()
            temp["이평선"] = f"MA{ma}"
            temp["이벤트"] = "상향돌파"
            logs.append(temp)

        down_log = x[x[f"DOWN_{ma}"]].copy()
        if not down_log.empty:
            temp = down_log[cols].copy()
            temp["이평선"] = f"MA{ma}"
            temp["이벤트"] = "하향이탈"
            logs.append(temp)

    if not logs:
        return pd.DataFrame()
    return pd.concat(logs, axis=0).sort_index()

# ---------------------------------------------------------
# 8) 오늘 상태
# ---------------------------------------------------------
def get_today_status_row(name, ticker, asset_type, x, total_tbl):
    last = x.iloc[-1].copy()
    best_ma = total_tbl.iloc[0]["MA"] if not total_tbl.empty and "MA" in total_tbl.columns else np.nan
    best_ma_name = f"MA{int(best_ma)}" if pd.notna(best_ma) else ""

    row = {
        "종목명": name,
        "티커": ticker,
        "분류": asset_type,
        "종가": safe_round(last.get("Close", np.nan), 2),
        "MA5": safe_round(last.get("MA5", np.nan), 2),
        "MA10": safe_round(last.get("MA10", np.nan), 2),
        "MA20": safe_round(last.get("MA20", np.nan), 2),
        "MA200": safe_round(last.get("MA200", np.nan), 2),
        "현재_MA5_위아래": get_above_below_text(last.get("Close"), last.get("MA5")),
        "현재_MA10_위아래": get_above_below_text(last.get("Close"), last.get("MA10")),
        "현재_MA20_위아래": get_above_below_text(last.get("Close"), last.get("MA20")),
        "현재_MA200_위아래": get_above_below_text(last.get("Close"), last.get("MA200")),
        "배열상태": get_stack_status_from_values(last.get("MA5"), last.get("MA10"), last.get("MA20"), last.get("MA200")),
        "거리_MA5(%)": safe_round(last.get("DIST_MA5", np.nan) * 100) if pd.notna(last.get("DIST_MA5", np.nan)) else np.nan,
        "거리_MA10(%)": safe_round(last.get("DIST_MA10", np.nan) * 100) if pd.notna(last.get("DIST_MA10", np.nan)) else np.nan,
        "거리_MA20(%)": safe_round(last.get("DIST_MA20", np.nan) * 100) if pd.notna(last.get("DIST_MA20", np.nan)) else np.nan,
        "거리_MA200(%)": safe_round(last.get("DIST_MA200", np.nan) * 100) if pd.notna(last.get("DIST_MA200", np.nan)) else np.nan,
        "최근거래대금": safe_round(last.get("Turnover", np.nan), 0),
        "20일평균거래대금": safe_round(last.get("Turnover20", np.nan), 0),
        "거래대금증가율_5vs20": safe_round(last.get("Turnover_Ratio_5_20", np.nan), 2),
        "가장유리한이평선": best_ma_name,
    }

    for ma in MA_LIST:
        row[f"오늘_상향돌파_MA{ma}"] = "Y" if bool(last.get(f"UP_{ma}", False)) else ""
        row[f"오늘_하향이탈_MA{ma}"] = "Y" if bool(last.get(f"DOWN_{ma}", False)) else ""

    return row

# ---------------------------------------------------------
# 9) 단일 실행
# ---------------------------------------------------------
def run_one(name, ticker, asset_type, start_date, end_date):
    df = download_price(ticker, start_date, end_date)
    if df.empty or len(df) < 300:
        print(f"[실패] {name} ({ticker}) 데이터 부족")
        return None

    turnover_filter = None if asset_type == "지수" else MIN_TURNOVER

    x = build_event_dataframe(
        df,
        ma_list=MA_LIST,
        forward_days=FORWARD_DAYS,
        cooldown=SIGNAL_COOLDOWN,
        min_turnover=turnover_filter
    )

    valid_years = list(range(CURRENT_YEAR - YEARS_BACK + 1, CURRENT_YEAR + 1))
    x_for_bt = x[x["Year"].isin(valid_years)].copy()
    if x_for_bt.empty:
        print(f"[실패] {name} ({ticker}) 유효 데이터 없음")
        return None

    total_tbl = summarize_total_by_ma(x_for_bt, MA_LIST, FORWARD_DAYS)
    year_tbl = summarize_yearly_by_ma(x_for_bt, MA_LIST, FORWARD_DAYS)
    event_log = build_event_log(x_for_bt, MA_LIST, FORWARD_DAYS)
    today_status = get_today_status_row(name, ticker, asset_type, x, total_tbl)

    return {
        "name": name,
        "ticker": ticker,
        "asset_type": asset_type,
        "raw": x,
        "bt_raw": x_for_bt,
        "total_table": total_tbl,
        "year_table": year_tbl,
        "event_log": event_log,
        "today_status": today_status,
    }

# ---------------------------------------------------------
# 10) 전체 실행
# ---------------------------------------------------------
def run_all():
    start_date = f"{CURRENT_YEAR - YEARS_BACK - 2}-01-01"
    end_date = TODAY
    results = {}

    print("=" * 90)
    print("최종 통합 시스템 시작")
    print("=" * 90)
    print(f"중복신호 제거: {SIGNAL_COOLDOWN}거래일")
    print(f"거래대금 필터: {MIN_TURNOVER:,.0f} 이상 (지수 제외)")
    print("=" * 90)

    for item in TARGETS:
        name = item["name"]
        ticker = item["ticker"]
        asset_type = item["asset_type"]
        print(f"분석 중 [{asset_type}] {name} ({ticker})")
        res = run_one(name, ticker, asset_type, start_date, end_date)
        if res is not None:
            results[name] = res

    return results

# ---------------------------------------------------------
# 11) 기본 점수
# ---------------------------------------------------------
def build_today_status_table(results):
    return pd.DataFrame([res["today_status"] for _, res in results.items()])

def build_rs_scores(results):
    rows = []
    for name, res in results.items():
        last = res["raw"].iloc[-1]
        rows.append({
            "종목명": name,
            "티커": res["ticker"],
            "분류": res["asset_type"],
            "RET_20": last.get("RET_20", np.nan),
            "RET_60": last.get("RET_60", np.nan),
            "RET_120": last.get("RET_120", np.nan),
            "RET_240": last.get("RET_240", np.nan),
        })
    df = pd.DataFrame(rows)
    for c in ["RET_20", "RET_60", "RET_120", "RET_240"]:
        df[f"{c}_pct"] = percentile_score(df[c])

    df["RS_SCORE"] = (
        df["RET_20_pct"] * RS_WEIGHT_20 +
        df["RET_60_pct"] * RS_WEIGHT_60 +
        df["RET_120_pct"] * RS_WEIGHT_120 +
        df["RET_240_pct"] * RS_WEIGHT_240
    )
    return df

def calc_ma_signal_score(today_status_df):
    df = today_status_df.copy()
    score = pd.Series(0.0, index=df.index)

    score += np.where(df["현재_MA5_위아래"] == "위", 8, 0)
    score += np.where(df["현재_MA10_위아래"] == "위", 14, 0)
    score += np.where(df["현재_MA20_위아래"] == "위", 26, 0)
    score += np.where(df["현재_MA200_위아래"] == "위", 20, 0)

    score += np.where(df["배열상태"] == "정배열", 20, 0)
    score += np.where(df["배열상태"] == "혼조", 5, 0)

    for ma in MA_LIST:
        score += np.where(df[f"오늘_상향돌파_MA{ma}"] == "Y", 8 if ma in [5, 10] else 12, 0)

    score += np.where((df["거리_MA20(%)"] > 0) & (df["거리_MA20(%)"] < 8), 8, 0)
    score += np.where((df["거리_MA200(%)"] > -5) & (df["거리_MA200(%)"] < 15), 4, 0)

    df["MA_SIGNAL_SCORE_RAW"] = score
    maxv = max(float(df["MA_SIGNAL_SCORE_RAW"].max()), 1.0)
    df["MA_SIGNAL_SCORE"] = (df["MA_SIGNAL_SCORE_RAW"] / maxv) * 100.0
    return df

def calc_turnover_score(today_status_df):
    df = today_status_df.copy()
    ratio = df["거래대금증가율_5vs20"].replace([np.inf, -np.inf], np.nan).fillna(0)

    score = []
    for v in ratio:
        if v >= 3.0:
            score.append(100)
        elif v >= 2.0:
            score.append(85)
        elif v >= 1.5:
            score.append(70)
        elif v >= 1.2:
            score.append(55)
        elif v >= 1.0:
            score.append(40)
        elif v >= 0.8:
            score.append(25)
        else:
            score.append(10)

    df["TURNOVER_SCORE"] = score
    return df

# ---------------------------------------------------------
# 12) RA 모멘텀
# ---------------------------------------------------------
def build_close_panel(results, only_universe=True):
    cols = {}
    for name, res in results.items():
        if only_universe and res["asset_type"] == "지수":
            continue
        s = res["raw"]["Close"].copy()
        s.name = name
        cols[name] = s
    if not cols:
        return pd.DataFrame()
    return pd.concat(cols.values(), axis=1).sort_index().dropna(how="all")

def calculate_risk_adjusted_momentum_table(close_df, period_1m=21, period_3m=63, period_6m=126, last_n_days=10):
    if close_df is None or close_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    daily_top_20 = {}
    daily_score_rows = []

    last_n_days = min(len(close_df), last_n_days)
    trading_days = close_df.index[-last_n_days:]

    for day in trading_days:
        end_idx = close_df.index.get_loc(day)
        start_1m = max(0, end_idx - period_1m + 1)
        start_3m = max(0, end_idx - period_3m + 1)
        start_6m = max(0, end_idx - period_6m + 1)

        win_1m = close_df.iloc[start_1m:end_idx + 1]
        win_3m = close_df.iloc[start_3m:end_idx + 1]
        win_6m = close_df.iloc[start_6m:end_idx + 1]

        ret_1m = win_1m.pct_change(fill_method=None).dropna().mean()
        ret_3m = win_3m.pct_change(fill_method=None).dropna().mean()
        ret_6m = win_6m.pct_change(fill_method=None).dropna().mean()

        vol_1m = win_1m.pct_change(fill_method=None).dropna().std()
        vol_3m = win_3m.pct_change(fill_method=None).dropna().std()
        vol_6m = win_6m.pct_change(fill_method=None).dropna().std()

        ra_1m = ret_1m / vol_1m
        ra_3m = ret_3m / vol_3m
        ra_6m = ret_6m / vol_6m

        avg_ra = ((ra_1m + ra_3m + ra_6m) / 3.0).replace([np.inf, -np.inf], np.nan).dropna()

        if not avg_ra.empty:
            top20 = avg_ra.nlargest(20)
            daily_top_20[day.strftime("%Y-%m-%d")] = top20.index.tolist()

            for name, score in avg_ra.items():
                daily_score_rows.append({
                    "날짜": day.strftime("%Y-%m-%d"),
                    "종목명": name,
                    "RA_1M": safe_round(ra_1m.get(name, np.nan), 4),
                    "RA_3M": safe_round(ra_3m.get(name, np.nan), 4),
                    "RA_6M": safe_round(ra_6m.get(name, np.nan), 4),
                    "RA_AVG": safe_round(score, 4),
                })

    top20_df = pd.DataFrame.from_dict(daily_top_20, orient="index", columns=[f"Top {i+1}" for i in range(20)])
    score_history_df = pd.DataFrame(daily_score_rows)
    return top20_df, score_history_df

def build_today_ra_rank(score_history_df):
    if score_history_df is None or score_history_df.empty:
        return pd.DataFrame()

    latest_date = score_history_df["날짜"].max()
    df = score_history_df[score_history_df["날짜"] == latest_date].copy()
    if df.empty:
        return pd.DataFrame()

    df = df.sort_values("RA_AVG", ascending=False).reset_index(drop=True)
    df.insert(0, "RA순위", range(1, len(df) + 1))
    df["RA_SCORE"] = percentile_score(df["RA_AVG"]).apply(lambda x: safe_round(x, 2))
    return df

# ---------------------------------------------------------
# 13) 시장 필터 + breadth
# ---------------------------------------------------------
def build_market_filter_info(results):
    today_df = build_today_status_table(results)
    market_df = today_df[today_df["분류"] == "지수"].copy()
    universe_df = today_df[today_df["분류"] != "지수"].copy()

    def get_market_row(name):
        x = market_df[market_df["종목명"] == name]
        return None if x.empty else x.iloc[0]

    kospi = get_market_row("KOSPI")
    kosdaq = get_market_row("KOSDAQ")

    info = {
        "market_score": 0,
        "market_view": "중립",
        "market_comment": "",
        "market_action": "중립",
        "ma20_above_count": 0,
        "ma20_above_ratio": np.nan,
        "ma200_above_count": 0,
        "ma200_above_ratio": np.nan,
        "stack_bull_count": 0,
        "stack_bull_ratio": np.nan,
        "universe_count": 0,
    }

    score = 0
    comments = []

    for label, row in [("KOSPI", kospi), ("KOSDAQ", kosdaq)]:
        if row is None:
            continue

        ma20 = row.get("현재_MA20_위아래", "")
        ma200 = row.get("현재_MA200_위아래", "")
        stack = row.get("배열상태", "")

        if ma20 == "위" and ma200 == "위" and stack == "정배열":
            score += MARKET_BONUS_STRONG
            comments.append(f"{label} 강세")
        elif ma20 == "위":
            score += MARKET_BONUS_NEUTRAL
            comments.append(f"{label} 단기양호")
        elif ma20 == "아래" and ma200 == "위":
            score += MARKET_PENALTY_WEAK
            comments.append(f"{label} 단기약세")
        elif ma20 == "아래" and ma200 == "아래":
            score += MARKET_PENALTY_VERY_WEAK
            comments.append(f"{label} 중기약세")

    if not universe_df.empty:
        universe_count = len(universe_df)

        ma20_above_count = int((universe_df["현재_MA20_위아래"] == "위").sum())
        ma200_above_count = int((universe_df["현재_MA200_위아래"] == "위").sum())
        stack_bull_count = int((universe_df["배열상태"] == "정배열").sum())

        ma20_above_ratio = ma20_above_count / universe_count * 100
        ma200_above_ratio = ma200_above_count / universe_count * 100
        stack_bull_ratio = stack_bull_count / universe_count * 100

        info["universe_count"] = universe_count
        info["ma20_above_count"] = ma20_above_count
        info["ma20_above_ratio"] = round(ma20_above_ratio, 1)
        info["ma200_above_count"] = ma200_above_count
        info["ma200_above_ratio"] = round(ma200_above_ratio, 1)
        info["stack_bull_count"] = stack_bull_count
        info["stack_bull_ratio"] = round(stack_bull_ratio, 1)

        if ma20_above_ratio >= 70:
            score += 10
            comments.append("유니버스 20일선 위 비율 높음")
        elif ma20_above_ratio >= 55:
            score += 5
            comments.append("유니버스 20일선 위 비율 양호")
        elif ma20_above_ratio < 35:
            score -= 10
            comments.append("유니버스 20일선 위 비율 낮음")
        elif ma20_above_ratio < 45:
            score -= 5
            comments.append("유니버스 20일선 위 비율 부진")

    info["market_score"] = int(score)
    info["market_comment"] = ", ".join(comments) if comments else "시장 데이터 부족"

    if score >= 20:
        info["market_view"] = "강한 우호"
        info["market_action"] = "적극매수 가능 구간"
    elif score >= 8:
        info["market_view"] = "우호적"
        info["market_action"] = "매수우위 대응"
    elif score >= -7:
        info["market_view"] = "중립"
        info["market_action"] = "선별 매수 / 추격매수 자제"
    elif score >= -19:
        info["market_view"] = "보수적"
        info["market_action"] = "방어적 대응 필요"
    else:
        info["market_view"] = "매우 보수적"
        info["market_action"] = "현금비중 확대 / 적극매수 자제"

    return info

# ---------------------------------------------------------
# 14) 코멘트 / 단계
# ---------------------------------------------------------
def generate_one_line_comment(row, market_info):
    parts = []

    if row.get("배열상태") == "정배열":
        parts.append("정배열")
    elif row.get("배열상태") == "역배열":
        parts.append("역배열")
    else:
        parts.append("혼조")

    parts.append("MA20 위" if row.get("현재_MA20_위아래") == "위" else "MA20 아래")
    parts.append("MA200 위" if row.get("현재_MA200_위아래") == "위" else "MA200 아래")

    if row.get("오늘_상향돌파_MA20") == "Y":
        parts.append("오늘 MA20 상향돌파")
    elif row.get("오늘_상향돌파_MA10") == "Y":
        parts.append("오늘 MA10 상향돌파")

    ratio = row.get("거래대금증가율_5vs20", np.nan)
    if pd.notna(ratio):
        if ratio >= 1.5:
            parts.append("거래대금 강증")
        elif ratio >= 1.1:
            parts.append("거래대금 증가")
        else:
            parts.append("거래대금 보통")

    ra_score = row.get("RA_SCORE", np.nan)
    if pd.notna(ra_score):
        if ra_score >= 80:
            parts.append("RA모멘텀 강함")
        elif ra_score >= 60:
            parts.append("RA모멘텀 양호")

    parts.append(f"시장:{market_info.get('market_view', '중립')}")

    if row.get("FINAL_SCORE", 0) >= 75:
        tail = "추세 지속형"
    elif row.get("FINAL_SCORE", 0) >= 60:
        tail = "관심 유지"
    else:
        tail = "보수 접근"

    return " + ".join(parts) + f" → {tail}"

def classify_stage(final_score, market_info, row):
    if pd.isna(final_score):
        return "보류"

    ma20_above = row.get("현재_MA20_위아래") == "위"
    ma200_above = row.get("현재_MA200_위아래") == "위"
    stack = row.get("배열상태") == "정배열"
    market_score = market_info.get("market_score", 0)

    if final_score >= 75 and ma20_above and (ma200_above or stack) and market_score > -20:
        return "매수"
    elif final_score >= 55 and ma20_above:
        return "관심"
    else:
        return "보류"

# ---------------------------------------------------------
# 15) 오늘 매수후보
# ---------------------------------------------------------
def build_buy_candidates_top10(results, today_ra_rank_df):
    today_df = build_today_status_table(results)
    rs_df = build_rs_scores(results)
    ma_df = calc_ma_signal_score(today_df)
    to_df = calc_turnover_score(today_df)
    market_info = build_market_filter_info(results)

    base = today_df.merge(
        rs_df[["종목명", "RS_SCORE", "RET_20", "RET_60", "RET_120", "RET_240"]],
        on="종목명",
        how="left"
    )
    base = base.merge(
        ma_df[["종목명", "MA_SIGNAL_SCORE", "MA_SIGNAL_SCORE_RAW"]],
        on="종목명",
        how="left"
    )
    base = base.merge(
        to_df[["종목명", "TURNOVER_SCORE"]],
        on="종목명",
        how="left"
    )

    if today_ra_rank_df is not None and not today_ra_rank_df.empty:
        base = base.merge(
            today_ra_rank_df[["종목명", "RA순위", "RA_AVG", "RA_SCORE"]],
            on="종목명",
            how="left"
        )
    else:
        base["RA순위"] = np.nan
        base["RA_AVG"] = np.nan
        base["RA_SCORE"] = np.nan

    base = base[base["분류"] == "유니버스"].copy()

    cond = (
        (base["현재_MA20_위아래"] == "위") |
        (base["오늘_상향돌파_MA20"] == "Y") |
        (base["오늘_상향돌파_MA10"] == "Y")
    )
    base = base[cond].copy()

    base["BASE_SCORE"] = (
        base["RS_SCORE"].fillna(0) * WEIGHT_RS +
        base["MA_SIGNAL_SCORE"].fillna(0) * WEIGHT_MA +
        base["TURNOVER_SCORE"].fillna(0) * WEIGHT_TURNOVER +
        base["RA_SCORE"].fillna(0) * WEIGHT_RA
    )

    base["MARKET_FILTER_SCORE"] = market_info["market_score"]
    base["FINAL_SCORE"] = base["BASE_SCORE"] + base["MARKET_FILTER_SCORE"]

    base["RET_20(%)"] = base["RET_20"] * 100
    base["RET_60(%)"] = base["RET_60"] * 100
    base["RET_120(%)"] = base["RET_120"] * 100
    base["RET_240(%)"] = base["RET_240"] * 100

    base["시장판단"] = market_info["market_view"]
    base["시장점수의미"] = market_info["market_action"]
    base["시장코멘트"] = market_info["market_comment"]

    base["매매단계"] = base.apply(lambda r: classify_stage(r["FINAL_SCORE"], market_info, r), axis=1)
    base["한줄코멘트"] = base.apply(lambda r: generate_one_line_comment(r, market_info), axis=1)

    stage_order = {"매수": 0, "관심": 1, "보류": 2}
    base["_stage_order"] = base["매매단계"].map(stage_order)
    base = base.sort_values(["_stage_order", "FINAL_SCORE"], ascending=[True, False]).reset_index(drop=True)
    base.drop(columns=["_stage_order"], inplace=True)

    base.insert(0, "순위", range(1, len(base) + 1))

    order_cols = [
        "순위", "종목명", "티커", "매매단계", "FINAL_SCORE", "BASE_SCORE", "MARKET_FILTER_SCORE",
        "RS_SCORE", "RA_SCORE", "RA순위", "RA_AVG", "MA_SIGNAL_SCORE", "TURNOVER_SCORE",
        "현재_MA5_위아래", "현재_MA10_위아래", "현재_MA20_위아래", "현재_MA200_위아래",
        "배열상태",
        "오늘_상향돌파_MA5", "오늘_상향돌파_MA10", "오늘_상향돌파_MA20", "오늘_상향돌파_MA200",
        "거래대금증가율_5vs20",
        "RET_20(%)", "RET_60(%)", "RET_120(%)", "RET_240(%)",
        "가장유리한이평선", "시장판단", "시장점수의미", "시장코멘트", "한줄코멘트"
    ]
    order_cols = [c for c in order_cols if c in base.columns]

    top10 = base[order_cols].head(10).reset_index(drop=True)
    buy_all = base[order_cols].copy()

    for df in [top10, buy_all]:
        if df is not None and not df.empty:
            num_cols = df.select_dtypes(include=[np.number]).columns
            for c in num_cols:
                df[c] = df[c].apply(lambda v: safe_round(v, 2) if pd.notna(v) else np.nan)

    return top10, buy_all, market_info

# ---------------------------------------------------------
# 16) 신규 진입
# ---------------------------------------------------------
def build_new_entry_candidates(buy_all):
    if buy_all is None or buy_all.empty:
        return pd.DataFrame(), pd.DataFrame()

    df = buy_all.copy()
    cond = (
        (df["오늘_상향돌파_MA20"] == "Y") |
        (df["오늘_상향돌파_MA10"] == "Y")
    )

    if STRICT_NEW_ENTRY:
        cond = cond & (df["현재_MA20_위아래"] == "위") & (df["거래대금증가율_5vs20"] >= 1.2)

    new_df = df[cond].copy()
    if new_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    new_df["신규진입강도"] = 0
    new_df.loc[new_df["오늘_상향돌파_MA10"] == "Y", "신규진입강도"] += 1
    new_df.loc[new_df["오늘_상향돌파_MA20"] == "Y", "신규진입강도"] += 2

    stage_order = {"매수": 0, "관심": 1, "보류": 2}
    new_df["_stage_order"] = new_df["매매단계"].map(stage_order).fillna(9)

    new_df = new_df.sort_values(["신규진입강도", "_stage_order", "FINAL_SCORE"], ascending=[False, True, False]).reset_index(drop=True)
    new_df.drop(columns=["_stage_order"], inplace=True)
    new_df.insert(0, "신규진입순위", range(1, len(new_df) + 1))

    show_cols = [
        "신규진입순위", "종목명", "티커", "매매단계",
        "FINAL_SCORE", "BASE_SCORE", "MARKET_FILTER_SCORE",
        "RS_SCORE", "RA_SCORE", "RA순위", "RA_AVG", "MA_SIGNAL_SCORE", "TURNOVER_SCORE",
        "오늘_상향돌파_MA10", "오늘_상향돌파_MA20",
        "현재_MA5_위아래", "현재_MA10_위아래", "현재_MA20_위아래", "현재_MA200_위아래",
        "배열상태", "거래대금증가율_5vs20",
        "RET_20(%)", "RET_60(%)", "RET_120(%)", "RET_240(%)",
        "가장유리한이평선", "시장판단", "시장점수의미", "시장코멘트", "한줄코멘트"
    ]
    show_cols = [c for c in show_cols if c in new_df.columns]

    return new_df[show_cols].head(10).reset_index(drop=True), new_df[show_cols].copy()

# ---------------------------------------------------------
# 17) 원페이지 요약
# ---------------------------------------------------------
def build_one_page_summary(results, buy_all=None):
    score_map, rank_map, stage_map, comment_map, ra_score_map, ra_rank_map = {}, {}, {}, {}, {}, {}

    if buy_all is not None and not buy_all.empty:
        for _, r in buy_all.iterrows():
            score_map[r["종목명"]] = r["FINAL_SCORE"]
            rank_map[r["종목명"]] = r["순위"]
            stage_map[r["종목명"]] = r["매매단계"]
            comment_map[r["종목명"]] = r["한줄코멘트"]
            ra_score_map[r["종목명"]] = r.get("RA_SCORE", np.nan)
            ra_rank_map[r["종목명"]] = r.get("RA순위", np.nan)

    rows = []
    for name, res in results.items():
        t = res["total_table"].copy()
        today = res["today_status"]
        if t.empty:
            continue

        best = t.iloc[0].copy()
        rows.append({
            "종목명": name,
            "티커": res["ticker"],
            "분류": res["asset_type"],
            "오늘매수순위": rank_map.get(name, ""),
            "오늘매수점수": score_map.get(name, np.nan),
            "매매단계": stage_map.get(name, ""),
            "RA순위": ra_rank_map.get(name, np.nan),
            "RA_SCORE": ra_score_map.get(name, np.nan),
            "가장유리한이평선": best.get("이평선", ""),
            "상향-하향_20일차이(%p)": best.get("상향-하향_20일평균차이(%p)", np.nan),
            "종합판단": best.get("종합판단", ""),
            "종가": today.get("종가", np.nan),
            "현재_MA5_위아래": today.get("현재_MA5_위아래", ""),
            "현재_MA10_위아래": today.get("현재_MA10_위아래", ""),
            "현재_MA20_위아래": today.get("현재_MA20_위아래", ""),
            "현재_MA200_위아래": today.get("현재_MA200_위아래", ""),
            "배열상태": today.get("배열상태", ""),
            "오늘_상향돌파_MA10": today.get("오늘_상향돌파_MA10", ""),
            "오늘_상향돌파_MA20": today.get("오늘_상향돌파_MA20", ""),
            "거래대금증가율_5vs20": today.get("거래대금증가율_5vs20", np.nan),
            "한줄코멘트": comment_map.get(name, ""),
        })

    out = pd.DataFrame(rows)
    if not out.empty:
        stage_order = {"매수": 0, "관심": 1, "보류": 2, "": 3}
        out["_stage_order"] = out["매매단계"].map(stage_order).fillna(3)
        out = out.sort_values(
            ["_stage_order", "오늘매수점수", "RA_SCORE", "상향-하향_20일차이(%p)"],
            ascending=[True, False, False, False],
            na_position="last"
        ).reset_index(drop=True)
        out.drop(columns=["_stage_order"], inplace=True)
    return out

def build_cross_total_summary(results):
    rows = []
    for name, res in results.items():
        t = res["total_table"].copy()
        if t.empty:
            continue
        t.insert(0, "종목명", name)
        t.insert(1, "티커", res["ticker"])
        t.insert(2, "분류", res["asset_type"])
        rows.append(t)
    return pd.concat(rows, axis=0, ignore_index=True) if rows else pd.DataFrame()

def build_cross_year_focus(results, focus_day=20):
    rows = []
    diff_col = f"상향-하향_{focus_day}일평균차이(%p)"
    for name, res in results.items():
        yt = res["year_table"].copy()
        if yt.empty or diff_col not in yt.columns:
            continue
        temp = yt[["Year", "이평선", "MA", diff_col]].copy()
        temp["종목명"] = name
        temp["티커"] = res["ticker"]
        temp["분류"] = res["asset_type"]
        rows.append(temp)

    if not rows:
        return pd.DataFrame()

    out = pd.concat(rows, axis=0, ignore_index=True)
    return out[["Year", "종목명", "티커", "분류", "이평선", "MA", diff_col]].sort_values(["Year", diff_col], ascending=[True, False]).reset_index(drop=True)

# ---------------------------------------------------------
# 18) 엑셀 포맷
# ---------------------------------------------------------
def autosize_worksheet(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len("" if cell.value is None else str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)

def apply_basic_excel_format(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    pos_fill = PatternFill("solid", fgColor="E2F0D9")
    neg_fill = PatternFill("solid", fgColor="FCE4D6")
    buy_fill = PatternFill("solid", fgColor="D9EAD3")
    watch_fill = PatternFill("solid", fgColor="FFF2CC")
    hold_fill = PatternFill("solid", fgColor="F4CCCC")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        autosize_worksheet(ws)
        header_map = {cell.column: cell.value for cell in ws[1]}

        for row in ws.iter_rows(min_row=2):
            stage_val = None
            for cell in row:
                cell.alignment = center_align
                if header_map.get(cell.column) == "매매단계":
                    stage_val = cell.value

            for cell in row:
                header_name = header_map.get(cell.column, "")
                if isinstance(cell.value, (int, float)) and isinstance(header_name, str):
                    if ("수익률" in header_name or "차이" in header_name or "점수" in header_name or "(%)" in header_name or "RA_" in header_name):
                        if cell.value > 0:
                            cell.fill = pos_fill
                        elif cell.value < 0:
                            cell.fill = neg_fill

            if stage_val == "매수":
                for cell in row:
                    if cell.fill.fill_type is None:
                        cell.fill = buy_fill
            elif stage_val == "관심":
                for cell in row:
                    if cell.fill.fill_type is None:
                        cell.fill = watch_fill
            elif stage_val == "보류":
                for cell in row:
                    if cell.fill.fill_type is None:
                        cell.fill = hold_fill

    wb.save(path)

# ---------------------------------------------------------
# 19) 엑셀 저장
# ---------------------------------------------------------
def save_to_excel(
    results,
    out_path,
    one_page,
    buy_top10,
    buy_all,
    new_entry_top10,
    new_entry_all,
    cross_total,
    cross_year,
    market_info,
    today_ra_rank_df,
    ra_top20_recent10d_df
):
    market_summary = pd.DataFrame([{
        "시장판단": market_info.get("market_view", ""),
        "시장점수": market_info.get("market_score", np.nan),
        "시장점수의미": market_info.get("market_action", ""),
        "시장코멘트": market_info.get("market_comment", ""),
        "유니버스종목수": market_info.get("universe_count", np.nan),
        "20일선위_종목수": market_info.get("ma20_above_count", np.nan),
        "20일선위_비율(%)": market_info.get("ma20_above_ratio", np.nan),
        "200일선위_종목수": market_info.get("ma200_above_count", np.nan),
        "200일선위_비율(%)": market_info.get("ma200_above_ratio", np.nan),
        "정배열_종목수": market_info.get("stack_bull_count", np.nan),
        "정배열_비율(%)": market_info.get("stack_bull_ratio", np.nan),
    }])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        market_summary.to_excel(writer, sheet_name="시장요약", index=False)

        if not one_page.empty:
            one_page.to_excel(writer, sheet_name="한페이지요약", index=False)
        if buy_top10 is not None and not buy_top10.empty:
            buy_top10.to_excel(writer, sheet_name="오늘매수후보_TOP10", index=False)
        if buy_all is not None and not buy_all.empty:
            buy_all.to_excel(writer, sheet_name="오늘매수후보_ALL", index=False)
        if new_entry_top10 is not None and not new_entry_top10.empty:
            new_entry_top10.to_excel(writer, sheet_name="오늘신규진입_TOP10", index=False)
        if new_entry_all is not None and not new_entry_all.empty:
            new_entry_all.to_excel(writer, sheet_name="오늘신규진입", index=False)
        if today_ra_rank_df is not None and not today_ra_rank_df.empty:
            today_ra_rank_df.to_excel(writer, sheet_name="오늘_RA모멘텀순위", index=False)
        if ra_top20_recent10d_df is not None and not ra_top20_recent10d_df.empty:
            ra_top20_recent10d_df.to_excel(writer, sheet_name="RA모멘텀_TOP20_최근10일", index=True)
        if not cross_total.empty:
            cross_total.to_excel(writer, sheet_name="전체_10년요약", index=False)
        if not cross_year.empty:
            cross_year.to_excel(writer, sheet_name="연도별_20일비교", index=False)

        for name, res in results.items():
            base = clip_sheet_title(name, 16)
            if not res["total_table"].empty:
                res["total_table"].to_excel(writer, sheet_name=clip_sheet_title(f"{base}_10년요약"), index=False)
            if not res["year_table"].empty:
                res["year_table"].to_excel(writer, sheet_name=clip_sheet_title(f"{base}_연도별"), index=False)
            if not res["event_log"].empty:
                res["event_log"].to_excel(writer, sheet_name=clip_sheet_title(f"{base}_이벤트로그"), index=True)

    apply_basic_excel_format(out_path)

# ---------------------------------------------------------
# 20) 구글시트 저장 - 요약 탭만 저장
# ---------------------------------------------------------
def get_gspread_client():
    try:
        from google.colab import auth
        auth.authenticate_user()

        import gspread
        from google.auth import default

        creds, _ = default()
        return gspread.authorize(creds)
    except Exception as e:
        print(f"[구글 인증 실패] {e}")
        return None

def write_df_to_ws(ws, df):
    ws.clear()
    if df is None or df.empty:
        ws.update("A1", [["데이터 없음"]])
        return

    df2 = df.copy()
    df2 = df2.replace([np.inf, -np.inf], np.nan).fillna("")

    for col in df2.columns:
        df2[col] = df2[col].astype(str).str.slice(0, 500)

    values = [df2.columns.tolist()] + df2.astype(str).values.tolist()
    ws.update(values, value_input_option="USER_ENTERED")

def safe_resize_ws(ws, rows=1000, cols=40):
    try:
        if ws.row_count != rows or ws.col_count != cols:
            ws.resize(rows=rows, cols=cols)
    except Exception as e:
        print(f"[시트 리사이즈 실패] {ws.title}: {e}")

def get_or_create_ws(sh, title, rows=1000, cols=40):
    titles = [w.title for w in sh.worksheets()]
    if title in titles:
        ws = sh.worksheet(title)
        safe_resize_ws(ws, rows=rows, cols=cols)
        return ws
    return sh.add_worksheet(title=title, rows=rows, cols=cols)

def trim_df_for_gsheet(df, max_rows=1000, max_cols=40):
    if df is None or df.empty:
        return df
    out = df.copy()
    if len(out) > max_rows:
        out = out.head(max_rows).copy()
    if out.shape[1] > max_cols:
        out = out.iloc[:, :max_cols].copy()
    return out

def save_to_gsheet(
    results,
    gsheet_name,
    one_page,
    buy_top10,
    buy_all,
    new_entry_top10,
    new_entry_all,
    cross_total,
    cross_year,
    market_info,
    today_ra_rank_df,
    ra_top20_recent10d_df
):
    gc = get_gspread_client()
    if gc is None:
        print("[구글시트 저장 실패] 인증 불가")
        return None

    try:
        sh = gc.open(gsheet_name)
    except Exception:
        sh = gc.create(gsheet_name)

    market_summary = pd.DataFrame([{
        "시장판단": market_info.get("market_view", ""),
        "시장점수": market_info.get("market_score", np.nan),
        "시장점수의미": market_info.get("market_action", ""),
        "시장코멘트": market_info.get("market_comment", ""),
        "유니버스종목수": market_info.get("universe_count", np.nan),
        "20일선위_종목수": market_info.get("ma20_above_count", np.nan),
        "20일선위_비율(%)": market_info.get("ma20_above_ratio", np.nan),
        "200일선위_종목수": market_info.get("ma200_above_count", np.nan),
        "200일선위_비율(%)": market_info.get("ma200_above_ratio", np.nan),
        "정배열_종목수": market_info.get("stack_bull_count", np.nan),
        "정배열_비율(%)": market_info.get("stack_bull_ratio", np.nan),
    }])

    sheet_specs = {
        "시장요약": (50, 15),
        "한페이지요약": (1500, 35),
        "오늘매수후보_TOP10": (50, 45),
        "오늘매수후보_ALL": (1500, 45),
        "오늘신규진입_TOP10": (50, 45),
        "오늘신규진입": (1000, 45),
        "오늘_RA모멘텀순위": (1000, 20),
        "RA모멘텀_TOP20_최근10일": (50, 25),
        "전체_10년요약": (1500, 40),
        "연도별_20일비교": (3000, 10),
    }

    dfs = {
        "시장요약": trim_df_for_gsheet(market_summary, 50, 15),
        "한페이지요약": trim_df_for_gsheet(one_page, 1500, 35),
        "오늘매수후보_TOP10": trim_df_for_gsheet(buy_top10, 50, 45),
        "오늘매수후보_ALL": trim_df_for_gsheet(buy_all, 1500, 45),
        "오늘신규진입_TOP10": trim_df_for_gsheet(new_entry_top10, 50, 45),
        "오늘신규진입": trim_df_for_gsheet(new_entry_all, 1000, 45),
        "오늘_RA모멘텀순위": trim_df_for_gsheet(today_ra_rank_df, 1000, 20),
        "RA모멘텀_TOP20_최근10일": trim_df_for_gsheet(ra_top20_recent10d_df, 50, 25),
        "전체_10년요약": trim_df_for_gsheet(cross_total, 1500, 40),
        "연도별_20일비교": trim_df_for_gsheet(cross_year, 3000, 10),
    }

    for title, df in dfs.items():
        rows, cols = sheet_specs[title]
        ws = get_or_create_ws(sh, title, rows=rows, cols=cols)
        write_df_to_ws(ws, df)

    print(f"[구글시트 저장 완료] {sh.url}")
    return sh.url

# ---------------------------------------------------------
# 21) 실행
